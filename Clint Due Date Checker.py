import pandas as pd
from datetime import *
from datetime import date
from tkinter import *
from tkinter import ttk, messagebox
import openpyxl
from configparser import ConfigParser
FILE = 'Client Due List.xlsx'

login_root = Tk()
login_root.title('Client Due List')
login_root.configure(bg='#BFD1DF')
logo_path = 'logo//instotech logo.png'
# adding logo
logo = PhotoImage(file=logo_path)
login_root.iconphoto(False, logo)
style = ttk.Style()
di = {'month': False,
	'selection': False}
# variable for counting the fail attempts while logging
error_count = 1
parser = ConfigParser()
parser.read("Client Due.ini")
users = dict(parser.items('username'))
pwds = dict(parser.items('password'))
user_pass = dict(zip(users.values(), pwds.values()))

# crating a frame in login window
login_frame = Frame(login_root, padx=10, pady=10)
login_frame.pack()

# label to display the specification about entry (username)
name_label = Label(login_frame, text='ENTER  YOUR  USERNAME', padx=10, pady=10)
name_label.grid(row=0, column=0)

# entry space for username
name = Entry(login_frame)
name.grid(row=0, column=2)

# label to display the specification about entry (password)
password_lab = Label(login_frame, text='PLEASE  ENTER  YOUR  PASSWORD', padx=10, pady=10, )
password_lab.grid(row=2, column=0)

login_fram = Frame(login_root, padx=10, pady=10)

# entry space for password
password = Entry(login_frame, show='*')
password.grid(row=2, column=2)

def month():
	di['month'] = True
	main()
	
def main():
	wb = openpyxl.load_workbook(FILE, read_only=False)
	ws = wb.active
	tables = ws.tables.values()

	tables_dict = {}

	NOW = datetime.now().date()
	due_list = []
	due_data_list = []
	if di['month']:
		month_sel = Toplevel()
		month_sel.title('Master Instruments Due')
		style = ttk.Style()

		search_cmb_bx = ttk.Combobox(month_sel, value=['Jan', 'Feb', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'], state='readonly')
		search_cmb_bx.current(0)
		search_cmb_bx.pack(padx=10, pady=10)

		def select():
			di['selection'] = search_cmb_bx.get()
			month_sel.destroy()

		but = Button(month_sel, text='Select', width=9, command=select)
		but.pack(padx=10, pady=10)
		month_sel.wait_window()
		for x in tables:
			if x.name == di['selection']:
				month_sel = x
				break
		tables_dict[month_sel.name] = {
		'table_name': month_sel.name,
		'worksheet': ws,
		'num_cols': len(month_sel.tableColumns),
		'table_range': month_sel.ref}
		data = ws[month_sel.ref]
		data_ = []
		rows_list = []
		for row in data:
			cols = []
			i = 0
			for i, col in enumerate(row):
				if col.value is None:
					if i < 5:
						cols.append(rows_list[-1][i])
					else:
						cols.append(' ')
					i += 1
				else:
					cols.append(col.value)
			rows_list.append(cols)

		df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

		li = []
		for i in range(len(df['NEXT CAL & DUEDATE'])):
			try:
				due_data_list.append(list(df.iloc[i]))
			except BaseException as e:
				print(e)
				print(i)
	else:
		for tbl in tables:
			tables_dict[tbl.name] = {
			'table_name': tbl.name,
			'worksheet': ws,
			'num_cols': len(tbl.tableColumns),
			'table_range': tbl.ref}
			data = ws[tbl.ref]
			data_ = []
			rows_list = []
			for row in data:
				cols = []
				i = 0
				for i, col in enumerate(row):
					if col.value is None:
						if i < 5:
							cols.append(rows_list[-1][i])
						else:
							cols.append(' ')
						i += 1
					else:
						cols.append(col.value)
				rows_list.append(cols)

			df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

			li = []
			for sn, due in enumerate(df['NEXT CAL & DUEDATE']):
				try:
					due = str(due)
					if due.__contains__('M'):
						due = due[due.index('M')+1:]
					if due.__contains__('Y'):
						due = due[due.index('Y')+1:]
					if due.__contains__('-'):
						due = due[due.index('-')+1:]
					if due.__contains__('-'):
						due = due[due.index('-')+1:]
					if due.__contains__('-'):
						due = due[due.index('-')+1:]
					if due.__contains__('�'):
						due = due[due.index('�')+1:]
					if due.__contains__('-'):
						due = due[due.index('-')+1:]
					if due.__contains__('Inst') and len(due) < 8:
						due = 'no_date'
					if due.__contains__('Inst'):
						due = due[due.index('Inst')+4:]
					if due.__contains__('dt.'):
						due = due[due.index('dt.')+3:]
					if due.strip().__contains__('to'):
						due = due[due.index('to')+2:]
					if due.strip().__contains__('&'):
						due = due[due.index('&')+1:]
					if due.__contains__('No Calibration') or due.__contains__('Not Calibrated') or due.__contains__('None'):
						due = 'no_date'
					if len(due.strip()) < 9:
						due = 'no_date'
					if due is None:
						due = 'no_date'
					if due.strip().__contains__(','):
						due = due.split(',')[-1]
					if due.strip() == '':
						due = 'no_date'
					li.append(due.lstrip().rstrip().strip())
					data = [due.lstrip().rstrip().strip(), str(sn.__add__(1)), tables_dict[tbl.name]['table_name'], ]
					if not data[0] == 'no_date' or data[0] == '':
						due_d, due_m, due_y = data[0].split('to')[-1].split('-')[-1].split('.')
						if due_m[0] == '0':
							due_m.replace('0', '')
						timedelta = NOW - date(int(due_y), int(due_m), int(due_d))
						if timedelta.days > -31 and timedelta.days < 31:
							due_data_list.append(list(df.iloc[int(data[1])-1]))
				except BaseException as e:
					print(e)
					print(due)
					pass
	login_fram.destroy()
	login_root.geometry('1300x500')
	login_root.configure(bg='#BFD1DF')
	# adding logo
	style = ttk.Style()

	style.theme_use('clam')

	# Configure the treeview Colours
	style.configure("Treeview", background='#FFFFFF', foreground='black', rowheight=25, fieldbackground="#FFFFFF")

	# Change selection colour
	style.map('Treeview',
			  background=[('selected', "#347083")])
	# Pick a theme
	style.configure("Treeview.Heading", font=(None, 12))

	try:
		for widget in login_root.winfo_children():
			widget.destroy()
		my_menu = Menu(login_root)
		login_root.config(menu=my_menu)

		company_menu = Menu(my_menu, tearoff=0)
		my_menu.add_cascade(label='Months', menu=company_menu)
		# Drop down menu
		company_menu.add_command(label='Select Month', command=month)
	except BaseException as e:
		print(e)
	tree_frame = LabelFrame(login_root, bg='#BFD1DF')
	tree_frame.grid(row=0, column=0)

	lab = Label(tree_frame, text=f"""The Following Clients Have Instruments { f'Due in this month:{di["selection"]}' if di['month'] else 'Due Or Are Due Within 30 Days'}""", font=('aerial', 16), fg='red', bg='#BFD1DF')
	lab.pack()
	tree_scroll_y = Scrollbar(tree_frame, orient='vertical')
	tree_scroll_y.pack(side=RIGHT, fill=Y)
	tree_scroll_x = Scrollbar(tree_frame, orient='horizontal')
	tree_scroll_x.pack(side=BOTTOM, fill=X)

	# Create the treeview
	tree_view = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set,
							selectmode="extended")
	tree_view.pack(padx=10, pady=10, fill=X)

	# Configure scroll bar
	tree_scroll_x.config(command=tree_view.xview)
	tree_scroll_y.config(command=tree_view.yview)

	# define columns
	tree_view["columns"] = ('sr', 'client', 'address', 'section', 'instruq', 'olcd', 'newcd', 'nameoftechni')
	# format our columns
	tree_view.column('#0', width=0, stretch=NO)
	tree_view.column('sr', anchor=W, width=15, minwidth=30)
	tree_view.column('client', anchor=W, width=150, minwidth=170)
	tree_view.column('address', anchor=W, width=190, minwidth=200)
	tree_view.column('section', anchor=W, width=190, minwidth=200)
	tree_view.column('instruq', anchor=W, width=190, minwidth=200)
	tree_view.column('olcd', anchor=W, width=200,minwidth=230)
	tree_view.column('newcd', anchor=W, width=200, minwidth=230)
	tree_view.column('nameoftechni', anchor=W, width=200, minwidth=220)

	# Create headings
	tree_view.heading('#0', text='', anchor=W)
	tree_view.heading('sr', text='SR. NO', anchor=W)
	tree_view.heading('client', text='CLIENTS  NAME', anchor=W)
	tree_view.heading('address', text='ADDRESS', anchor=W)
	tree_view.heading('section', text='SECTION', anchor=W)
	tree_view.heading('instruq', text='INSTRUMENT QTY.', anchor=W)
	tree_view.heading('olcd', text='OLD CAL & DUEDATE', anchor=W)
	tree_view.heading('newcd', text='NEXT CAL & DUEDATE', anchor=W)
	tree_view.heading('nameoftechni', text='NAME OF TECHNICIAN', anchor=W)

	tree_view.tag_configure('oddrow', background='#84CCDC', font=('aerial', 14))
	tree_view.tag_configure('evenrow', background='#FFF44F', font=('aerial', 14))

	for record in tree_view.get_children():
		tree_view.delete(record)
	if len(due_data_list) > 0:
		for ind,row in enumerate(due_data_list):
			try:
				if int(ind) % 2 == 0:
					tree_view.insert(parent='', index='end', iid=ind, values=list(row), tags=('evenrow',))
				else:
					tree_view.insert(parent='', index='end', iid=ind, values=list(row), tags=('oddrow',))
			except BaseException as e:
				pass	
	else:
		messagebox.showinfo('No Due Instruments', f"""NO Clients Were Found That have Instruments {f'Due in this month:{di["selection"]}' if di['month'] else 'due Or to be Due in 30 Days'}""")

def log_check():
	user_name_check = name.get()
	if name.get() in user_pass and user_pass[name.get()] == password.get():
		try:
			login_frame.destroy()
		except BaseException as e:
			print(e)
		login_fram.grid(row=0, column=0)
		login_root.geometry('450x300')

		lab = Label(login_fram, text='What Would You Like to do?', font=('aerial', 15))
		lab.grid(row=0, column=0)
		all_months = Button(login_fram, text='Normal Search', font=('aerial', 14), width=12, height=6, command=main)
		all_months.grid(row=1, column=0)
		only_months = Button(login_fram, text='Monthly Search', font=('aerial', 14), width=12, height=6, command=month)
		only_months.grid(row=1, column=1)
		

	else:
		global error_count
		name.delete(0, END)
		password.delete(0, END)
		# loop to prevent many no of labels saying invalid username or id
		if error_count == 1:
			label_error = Label(login_frame, text='Invalid username or Password', padx=10, pady=10, fg='red')
			label_error.grid(row=3, column=0)
			error_count += 1
# button to start check command
check_button = Button(login_frame, text='ENTER', command=log_check)
check_button.grid(row=3, column=1)
cancel_button = Button(login_frame, text='CANCEL', command=login_frame.quit)
cancel_button.grid(row=3, column=2)
login_root.mainloop()
